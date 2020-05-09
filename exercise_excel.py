'''
每行都要写注释，直接提交代码

使用openpyxl实现以下需求
1.使用excel 写入一组数据，姓名，身高，体重
2.计算是否为健康体重，如果是健康体重，则在旁边备注健康，并将姓名打印出来
•健康体重计算公式：（身高cm-70）×60%

(可以做一部分优化)

'''
# 从openpyxl里导入Workbook和load_workbook类
from openpyxl import Workbook, load_workbook

# 声明TestExcel类
class TestExcel:
    # 初始化，给文件名、页签名、文件里的姓名，身高和体重赋值
    # 参数name，提示类型是list
    # 参数height_weight，提示类型是dict
    def __init__(self, file_name, sheet_name, name:list, height_weight:dict):
        # self调用类的属性，用init传进来的参数初始化
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.name = name
        self.height_weight = height_weight

    # 定义创建文件的方法
    def create_data(self):
        # 对Workbook进行实例化
        work_book = Workbook()
        # 用Workbook实例获取当前活跃页签，并赋值给变量work_sheet
        work_sheet = work_book.active
        # self调用类属性sheet_name，给当前活跃页签命名
        work_sheet.title = self.sheet_name
        # 定义dict类型的变量 column_dict，用于保存标题
        column_dict = {"A1":"姓名", "B1":"身高", "C1":"体重", "D1":"备注"}
        # 定义list类型的变量column_keys，用于保存单元格位置
        # 用列表推导式，将单元格位置保存到变量column_keys，用keys()可以取得字典中的key值(单元格位置)
        column_keys = [i for i in column_dict.keys()]
        # 定义list类型的变量height_weight_keys，用于保存身高
        # 用列表推导式，将身高保存到变量height_weight_keys，用keys()可以取得字典中的key值(身高)
        height_weight_keys = [i for i in self.height_weight.keys()]

        # 用for循环实现4个标题的插入，range(4)表示0，1，2和3次循环
        for i in range(4):
            # column_keys[i]表示标题插入的单元格位置
            # column_dict[column_keys[i]]表示标题插入的内容
            work_sheet[column_keys[i]] = column_dict[column_keys[i]]

        # 用for循环实现姓名、身高和体重的插入，len(self.name)表示取类属性name的长度
        for i in range(len(self.name)):
            # 从第二行开始，给第一列单元格赋值姓名，self调用类属性name
            work_sheet.cell(row=i+2, column=1).value = self.name[i]
            # 从第二行开始，给第二列单元格赋值身高，height_weight_keys[i]是身高（取key值）
            work_sheet.cell(row=i+2, column=2).value = height_weight_keys[i]
            # 从第二行开始，给第三列单元格赋值体重，height_weight[height_weight_keys[i]是体重（用key值取value）
            work_sheet.cell(row=i+2, column=3).value = self.height_weight[height_weight_keys[i]]

        # save()保存文件
        work_book.save(self.file_name)

    # 定义获得健康体重的方法
    def get_health(self):
        # 对load_workbook进行实例化，通过文件名读取文件，文件名是用self调用类的属性file_name
        load_excel = load_workbook(filename=self.file_name)
        # 通过页签名读取页签，页签名是用self调用类的属性sheet_name获得，赋值给变量load_sheet
        load_sheet = load_excel[self.sheet_name]

        # 用for循环实现姓名、身高和体重的读取，len(self.name)表示取类属性name的长度
        for i in range(len(self.name)):
            # 从第二行开始，读取第一列单元格中的姓名，赋值给变量name
            name = load_sheet.cell(row=i+2, column=1).value
            # 从第二行开始，读取第二列单元格中的身高，赋值给变量height
            height = load_sheet.cell(row=i+2, column=2).value
            # 从第二行开始，读取第三列单元格中的体重，赋值给变量weight
            weight = load_sheet.cell(row=i+2, column=3).value
            # 定义变量health_weight，用于计算健康体重
            health_weight = (height-70)*0.6

            # 如果文件中的体重与它对应算出来的健康体重相等，则给第四列单元格赋值他的姓名，并标注“是健康体重”
            if weight == health_weight:
                load_sheet.cell(row=i+2, column=4).value = name+"是健康体重"
                # 打印出姓名，用逗号分隔，逗号会打印出空格
                print(name, "是健康体重")
                # 打印出姓名，用{}占位符和format打印，就没有空格了
                # print("{}是健康体重".format(name))

        # save()保存文件
        load_excel.save(self.file_name)

# 定义文件名变量，并赋值
excel_file_name = "test.xlsx"
# 定义页签名变量，并赋值
excel_sheet_name = "statistics"
# 定义存放姓名的list变量，并赋值
name_list = ["张三", "李四", "王五", "赵六", "孙七", "周八"]
# 定义存放身高和体重的dict变量，并赋值
height_weight_dict = {170:60, 180:70, 160:54, 185:90, 175:130, 155:40}

# 声明TestExcel类的实例，并初始化文件名、页签名、文件里的姓名，身高和体重
excel = TestExcel(excel_file_name, excel_sheet_name, name_list, height_weight_dict)
# 用TestExcel类的实例调用创建文件的方法
excel.create_data()
# 用TestExcel类的实例调用获得健康体重的方法
excel.get_health()
